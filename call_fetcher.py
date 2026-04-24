"""
call_fetcher.py — Horizon Europe Çağrı Çekici
EC API (POST) + Euresearch + UfukAvrupa + Yerel DB
"""
import requests
import time
import re
import hashlib
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from io import BytesIO

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════
# CACHE
# ═══════════════════════════════════════════════════════════
class CallCache:
    def __init__(self, ttl_minutes: int = 30):
        self.ttl = ttl_minutes * 60
        self._store: Dict[str, dict] = {}

    def get(self, key: str):
        entry = self._store.get(key)
        if entry and (time.time() - entry["ts"]) < self.ttl:
            return entry["data"]
        return None

    def set(self, key: str, data):
        self._store[key] = {"data": data, "ts": time.time()}

    def clear(self):
        self._store.clear()


# ═══════════════════════════════════════════════════════════
# HTML TEMİZLEYİCİ
# ═══════════════════════════════════════════════════════════
def clean_html(text) -> str:
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    if not text:
        return ""
    if HAS_BS4:
        try:
            soup = BeautifulSoup(text, "html.parser")
            return soup.get_text(separator=" ", strip=True)
        except Exception:
            pass
    text = re.sub(r'<br\s*/?>', ' ', text)
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&quot;', '"', text)
    text = re.sub(r'&#39;', "'", text)
    text = re.sub(r'&#\d+;', '', text)
    text = re.sub(r'&\w+;', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# ═══════════════════════════════════════════════════════════
# EC API HELPERS
# ═══════════════════════════════════════════════════════════
EC_SEARCH_URL = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"


def _safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, list):
        parts = []
        for v in val:
            if isinstance(v, dict):
                parts.append(str(v.get("value", v.get("code", ""))))
            else:
                parts.append(str(v))
        return ", ".join(parts)
    if isinstance(val, dict):
        return str(val.get("value", val.get("code", "")))
    return str(val)


def _extract_ec_field(metadata: dict, *field_names) -> str:
    for target in field_names:
        target_lower = target.lower()
        for key, values in metadata.items():
            if target_lower in key.lower():
                return clean_html(_safe_str(values))
    return ""


def _extract_ec_list(metadata: dict, *field_names) -> List[str]:
    for target in field_names:
        target_lower = target.lower()
        for key, values in metadata.items():
            if target_lower in key.lower():
                if isinstance(values, list):
                    return [clean_html(_safe_str(v)) for v in values if v]
                return [clean_html(_safe_str(values))]
    return []


def _normalize_status(raw: str) -> str:
    if not raw:
        return "Open"
    low = raw.lower()
    if any(x in low for x in ["open", "31094501"]):
        return "Open"
    if any(x in low for x in ["forthcoming", "31094502", "upcoming"]):
        return "Forthcoming"
    if any(x in low for x in ["closed", "31094503"]):
        return "Closed"
    return raw


def _detect_action_types_from_text(call_id: str, title: str) -> List[str]:
    combined = (call_id + " " + title).upper()
    types = []
    patterns = [
        ("ERC-StG", r'ERC.{0,10}(STG|STARTING)'),
        ("ERC-CoG", r'ERC.{0,10}(COG|CONSOLIDATOR)'),
        ("ERC-AdG", r'ERC.{0,10}(ADG|ADVANCED)'),
        ("ERC-PoC", r'ERC.{0,10}(POC|PROOF)'),
        ("MSCA-DN", r'MSCA.{0,10}(DN|DOCTORAL)'),
        ("MSCA-PF", r'MSCA.{0,10}(PF|POSTDOC)'),
        ("EIC-Pathfinder", r'EIC.{0,10}PATHFINDER'),
        ("EIC-Accelerator", r'EIC.{0,10}ACCELERATOR'),
        ("CSA", r'\bCSA\b'),
        ("IA", r'\bIA\b'),
        ("RIA", r'\bRIA\b'),
    ]
    for name, pattern in patterns:
        if re.search(pattern, combined):
            types.append(name)
    return types if types else ["RIA"]


def _parse_ec_result(meta: dict) -> Optional[Dict]:
    try:
        call_id = _extract_ec_field(meta, "identifier", "callIdentifier", "topicIdentifier")
        title = _extract_ec_field(meta, "title")

        if not call_id and not title:
            return None

        status_raw = _extract_ec_field(meta, "status")
        deadline = _extract_ec_field(meta, "deadlineDate", "deadlineDates")
        start_date = _extract_ec_field(meta, "startDate", "openingDate")
        budget = _extract_ec_field(meta, "budget", "budgetOverviewReference")
        desc = _extract_ec_field(meta, "description", "smedescription")
        programme = _extract_ec_field(meta, "programmeDivision", "destination")
        types_list = _extract_ec_list(meta, "typesOfAction", "typeOfAction")
        keywords_list = _extract_ec_list(meta, "keywords", "tags")

        status_display = _normalize_status(status_raw)

        if deadline:
            date_match = re.search(r'(\d{4}-\d{2}-\d{2})', deadline)
            if date_match:
                deadline = date_match.group(1)
            else:
                deadline = deadline[:10] if len(deadline) >= 10 else "N/A"

        if not types_list:
            types_list = _detect_action_types_from_text(call_id, title)

        link = ""
        if call_id:
            clean_id = call_id.split(",")[0].strip()
            link = (
                f"https://ec.europa.eu/info/funding-tenders/"
                f"opportunities/portal/screen/opportunities/"
                f"topic-details/{clean_id}"
            )

        return {
            "call_id": call_id or "N/A",
            "title": title or call_id or "N/A",
            "status": status_display,
            "deadline": deadline if deadline else "N/A",
            "start_date": start_date[:10] if start_date else "",
            "budget_total": budget,
            "budget_per_project": "",
            "action_types": types_list if types_list else ["RIA"],
            "description": desc[:500] if desc else "",
            "keywords": keywords_list,
            "destination": programme,
            "scope": desc[:1000] if desc else "",
            "expected_outcomes": "",
            "link": link,
            "source": "EC API",
            "topics": [{"topic_id": call_id.split(",")[0].strip()}] if call_id else [],
        }
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════
# EC API — POST + GET FALLBACK + PAGINATION
# ═══════════════════════════════════════════════════════════
def fetch_horizon_calls(
    search_text: str = "",
    status: str = "",
    max_results: int = 500,
    max_pages: int = 5,
) -> Tuple[List[Dict], dict]:
    """
    EC Funding & Tenders API — Çoklu endpoint + format denemesi.
    """
    calls = []
    debug_info = {
        "attempts": [],
        "success": False,
        "total_api": 0,
        "pages_fetched": 0,
    }

    # ─── Strateji 1: JSON bool query POST ───
    try:
        url1 = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"
        body1 = {
            "apiKey": "SEDIA",
            "text": search_text if search_text else "*",
            "pageSize": min(max_results, 100),
            "pageNumber": 1,
            "sort": "deadlineDate:desc",
            "query": {
                "bool": {
                    "must": [
                        {"term": {"programmePeriod": "2021-2027"}},
                        {"term": {"frameworkProgramme": "43108390"}},
                    ]
                }
            },
        }
        if status:
            sm = {"open": "31094501", "forthcoming": "31094502", "closed": "31094503"}
            sc = sm.get(status.lower(), "")
            if sc:
                body1["query"]["bool"]["must"].append({"term": {"status": sc}})

        resp1 = requests.post(
            url1,
            headers={"Content-Type": "application/json", "Accept": "application/json"},
            json=body1,
            timeout=25,
        )
        debug_info["attempts"].append({
            "strategy": "rest_post_json_query",
            "url": url1,
            "status": resp1.status_code,
            "body_type": "json_bool_query",
        })
        if resp1.status_code == 200:
            data = resp1.json()
            for r in data.get("results", []):
                c = _parse_ec_result(r.get("metadata", {}))
                if c:
                    calls.append(c)
            debug_info["attempts"][-1]["count"] = len(calls)
    except Exception as e:
        debug_info["attempts"].append({
            "strategy": "rest_post_json_query",
            "error": str(e)[:200],
        })

    # ─── Strateji 2: String query POST + SAYFALAMA (ÇALIŞIYOR!) ───
    if not calls:
        try:
            url2 = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"
            page_size = min(max_results, 100)

            for page in range(1, max_pages + 1):
                params_str = (
                    f"apiKey=SEDIA"
                    f"&text={'horizon ' + search_text if search_text else '*'}"
                    f"&pageSize={page_size}"
                    f"&pageNumber={page}"
                    f"&query=programmePeriod%2Fcode%3D%272021-2027%27"
                    f"%20AND%20frameworkProgramme%2Fcode%3D%2743108390%27"
                )

                # Status filtresi
                if status:
                    sm = {
                        "open": "31094501",
                        "forthcoming": "31094502",
                        "closed": "31094503",
                    }
                    sc = sm.get(status.lower(), "")
                    if sc:
                        params_str += f"%20AND%20status%2Fcode%3D%27{sc}%27"

                resp2 = requests.post(
                    f"{url2}?{params_str}",
                    headers={"Accept": "application/json"},
                    timeout=25,
                )

                if page == 1:
                    debug_info["attempts"].append({
                        "strategy": "rest_post_query_string",
                        "status": resp2.status_code,
                    })

                if resp2.status_code != 200:
                    break

                data = resp2.json()
                total_results = data.get("totalResults", 0)
                results = data.get("results", [])

                if page == 1:
                    debug_info["attempts"][-1]["total_results"] = total_results
                    debug_info["attempts"][-1]["page1_count"] = len(results)

                if not results:
                    break

                for r in results:
                    c = _parse_ec_result(r.get("metadata", {}))
                    if c:
                        calls.append(c)

                debug_info["pages_fetched"] = page

                if len(calls) >= max_results:
                    break
                if len(results) < page_size:
                    break

                time.sleep(0.3)

            if debug_info["attempts"] and debug_info["attempts"][-1].get("strategy") == "rest_post_query_string":
                debug_info["attempts"][-1]["count"] = len(calls)

        except Exception as e:
            debug_info["attempts"].append({
                "strategy": "rest_post_query_string",
                "error": str(e)[:200],
            })

    # ─── Strateji 3: Portal Topic Search API ───
    if not calls:
        try:
            url3 = (
                "https://ec.europa.eu/info/funding-tenders/opportunities/"
                "data/topicSearch/search.json"
            )
            params3 = {
                "callStatus": status.upper() if status else "OPEN,FORTHCOMING",
                "frameworkProgramme": "HORIZON",
                "pageSize": str(min(max_results, 100)),
                "pageNumber": "0",
                "sortBy": "deadlineDate",
                "orderBy": "desc",
            }
            if search_text:
                params3["keywords"] = search_text

            resp3 = requests.get(
                url3,
                params=params3,
                headers={
                    "Accept": "application/json",
                    "User-Agent": "GrantMirror-AI/1.0",
                },
                timeout=25,
            )
            debug_info["attempts"].append({
                "strategy": "portal_topic_search",
                "url": url3,
                "status": resp3.status_code,
            })
            if resp3.status_code == 200:
                data = resp3.json()
                topics = data.get("topicResults", data.get("results", []))
                if isinstance(topics, dict):
                    topics = topics.get("content", topics.get("topics", []))
                for t in topics:
                    call = _parse_portal_topic(t)
                    if call:
                        calls.append(call)
                debug_info["attempts"][-1]["count"] = len(calls)
        except Exception as e:
            debug_info["attempts"].append({
                "strategy": "portal_topic_search",
                "error": str(e)[:200],
            })

    # ─── Strateji 4: Form-encoded POST ───
    if not calls:
        try:
            url4 = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"
            form_data = {
                "apiKey": "SEDIA",
                "text": search_text if search_text else "*",
                "pageSize": str(min(max_results, 100)),
                "pageNumber": "1",
                "query": (
                    "programmePeriod/code='2021-2027' AND "
                    "frameworkProgramme/code='43108390'"
                ),
            }
            resp4 = requests.post(
                url4,
                data=form_data,
                headers={"Accept": "application/json"},
                timeout=25,
            )
            debug_info["attempts"].append({
                "strategy": "rest_post_form_encoded",
                "status": resp4.status_code,
            })
            if resp4.status_code == 200:
                data = resp4.json()
                for r in data.get("results", []):
                    c = _parse_ec_result(r.get("metadata", {}))
                    if c:
                        calls.append(c)
                debug_info["attempts"][-1]["count"] = len(calls)
        except Exception as e:
            debug_info["attempts"].append({
                "strategy": "rest_post_form_encoded",
                "error": str(e)[:200],
            })

    # ─── Strateji 5: EC Portal HTML scrape (son çare) ───
    if not calls:
        try:
            calls_scrape = _scrape_ec_portal(search_text, status, max_results)
            calls.extend(calls_scrape)
            debug_info["attempts"].append({
                "strategy": "ec_portal_scrape",
                "count": len(calls_scrape),
            })
        except Exception as e:
            debug_info["attempts"].append({
                "strategy": "ec_portal_scrape",
                "error": str(e)[:200],
            })

    # ─── Dedup ───
    seen = set()
    unique = []
    for c in calls:
        cid = c.get("call_id", "")
        if cid and cid not in seen:
            seen.add(cid)
            unique.append(c)
        elif not cid:
            unique.append(c)

    debug_info["total_api"] = len(unique)
    debug_info["before_dedup"] = len(calls)
    debug_info["success"] = len(unique) > 0
    if unique:
        debug_info["winning_strategy"] = next(
            (a["strategy"] for a in debug_info["attempts"] if a.get("count", 0) > 0),
            "unknown",
        )

    return unique[:max_results], debug_info


def _parse_portal_topic(topic: dict) -> Optional[Dict]:
    """EC Portal topic search sonucunu parse et."""
    try:
        call_id = topic.get("identifier", topic.get("topicIdentifier", ""))
        title = topic.get("title", topic.get("topicTitle", ""))
        status = topic.get("callStatus", topic.get("status", "Open"))
        deadline = topic.get("deadlineDate", topic.get("deadline", ""))
        budget = topic.get("budgetOverall", topic.get("budget", ""))
        desc = topic.get("description", topic.get("shortDescription", ""))
        action_types = topic.get("typesOfAction", [])

        if isinstance(action_types, str):
            action_types = [action_types]

        if not call_id and not title:
            return None

        # Status normalleştir
        if isinstance(status, str):
            status = _normalize_status(status)

        # Deadline normalleştir
        if deadline:
            dm = re.search(r'(\d{4}-\d{2}-\d{2})', str(deadline))
            if dm:
                deadline = dm.group(1)
            else:
                deadline = str(deadline)[:10]

        link = ""
        if call_id:
            clean_id = str(call_id).split(",")[0].strip()
            link = (
                f"https://ec.europa.eu/info/funding-tenders/"
                f"opportunities/portal/screen/opportunities/"
                f"topic-details/{clean_id}"
            )

        return {
            "call_id": str(call_id) or "N/A",
            "title": clean_html(str(title)) or str(call_id) or "N/A",
            "status": status,
            "deadline": deadline if deadline else "N/A",
            "start_date": "",
            "budget_total": str(budget) if budget else "",
            "budget_per_project": "",
            "action_types": action_types if action_types else ["RIA"],
            "description": clean_html(str(desc))[:500],
            "keywords": [],
            "destination": "",
            "scope": clean_html(str(desc))[:1000],
            "expected_outcomes": "",
            "link": link,
            "source": "EC Portal",
            "topics": [{"topic_id": str(call_id)}] if call_id else [],
        }
    except Exception:
        return None


def _parse_cordis_result(result: dict) -> Optional[Dict]:
    """CORDIS API sonucunu parse et."""
    try:
        title = result.get("title", "")
        acronym = result.get("acronym", "")
        call_id = result.get("callIdentifier", result.get("id", ""))
        status = result.get("status", "Open")
        start_date = result.get("startDate", "")
        end_date = result.get("endDate", "")

        if not title:
            return None

        return {
            "call_id": str(call_id) or f"CORDIS-{hashlib.md5(title.encode()).hexdigest()[:8]}",
            "title": clean_html(f"{acronym}: {title}" if acronym else title),
            "status": _normalize_status(str(status)),
            "deadline": end_date[:10] if end_date else "N/A",
            "start_date": start_date[:10] if start_date else "",
            "budget_total": "",
            "budget_per_project": "",
            "action_types": ["RIA"],
            "description": clean_html(result.get("objective", ""))[:500],
            "keywords": [],
            "destination": "",
            "scope": clean_html(result.get("objective", ""))[:1000],
            "expected_outcomes": "",
            "link": f"https://cordis.europa.eu/project/id/{call_id}" if call_id else "",
            "source": "CORDIS",
            "topics": [],
        }
    except Exception:
        return None


def _scrape_ec_portal(
    search_text: str = "",
    status: str = "",
    max_results: int = 50,
) -> List[Dict]:
    """EC F&T Portal'ı doğrudan scrape et (son çare)."""
    if not HAS_BS4:
        return []

    calls = []
    try:
        url = (
            "https://ec.europa.eu/info/funding-tenders/opportunities/"
            "portal/screen/opportunities/topic-search"
        )
        params = {
            "programmePart": "HORIZON",
            "callStatus": status.upper() if status else "OPEN,FORTHCOMING",
        }
        if search_text:
            params["keywords"] = search_text

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
            ),
            "Accept": "text/html,application/json",
        }

        resp = requests.get(url, params=params, headers=headers, timeout=25)
        if resp.status_code != 200:
            return []

        # JSON yanıt mı kontrol et
        try:
            data = resp.json()
            topics = (
                data.get("topicResults", {}).get("content", [])
                if isinstance(data.get("topicResults"), dict)
                else data.get("results", [])
            )
            for t in topics[:max_results]:
                call = _parse_portal_topic(t)
                if call:
                    calls.append(call)
            return calls
        except (ValueError, KeyError):
            pass

        # HTML yanıtsa scrape et
        soup = BeautifulSoup(resp.text, "html.parser")
        seen = set()
        for el in soup.find_all(["div", "tr", "article"]):
            text = el.get_text(separator=" ", strip=True)
            if len(text) < 20:
                continue
            if not re.search(r'HORIZON|ERC|EIC|MSCA', text, re.IGNORECASE):
                continue

            title = ""
            for h in el.find_all(["h3", "h4", "a", "strong"]):
                t = clean_html(h.get_text(strip=True))
                if len(t) >= 10:
                    title = t
                    break
            if not title or title.lower()[:40] in seen:
                continue
            seen.add(title.lower()[:40])

            link = ""
            a = el.find("a", href=True)
            if a:
                link = a.get("href", "")
                if link and not link.startswith("http"):
                    link = f"https://ec.europa.eu{link}"

            deadline = "N/A"
            dm = re.search(r'(\d{4}-\d{2}-\d{2})', text)
            if dm:
                deadline = dm.group(1)

            calls.append({
                "call_id": f"EC-{hashlib.md5(title.encode()).hexdigest()[:8]}",
                "title": title,
                "status": "Open",
                "deadline": deadline,
                "start_date": "",
                "budget_total": "",
                "budget_per_project": "",
                "action_types": _detect_action_types_from_text("", title),
                "description": clean_html(text)[:500],
                "keywords": [],
                "destination": "",
                "scope": clean_html(text)[:1000],
                "expected_outcomes": "",
                "link": link,
                "source": "EC Scrape",
                "topics": [],
            })

            if len(calls) >= max_results:
                break

    except Exception:
        pass

    return calls

# ═══════════════════════════════════════════════════════════
# TOPIC DETAILS
# ═══════════════════════════════════════════════════════════
def fetch_topic_details(topic_id: str) -> Optional[Dict]:
    if not topic_id or topic_id == "N/A":
        return None
    try:
        body = {
            "apiKey": "SEDIA",
            "text": f'"{topic_id}"',
            "pageSize": 5,
            "pageNumber": 1,
        }
        headers = {"Content-Type": "application/json", "Accept": "application/json"}

        resp = requests.post(EC_SEARCH_URL, headers=headers, json=body, timeout=15)
        if resp.status_code != 200:
            resp = requests.get(
                EC_SEARCH_URL,
                params={k: str(v) for k, v in body.items()},
                timeout=15,
            )

        if resp.status_code == 200:
            data = resp.json()
            for r in data.get("results", []):
                meta = r.get("metadata", {})
                found_id = _extract_ec_field(meta, "identifier")
                if topic_id.lower() in found_id.lower():
                    return {
                        "topic_id": topic_id,
                        "title": _extract_ec_field(meta, "title"),
                        "description": _extract_ec_field(meta, "description"),
                        "budget": _extract_ec_field(meta, "budget"),
                        "deadline": _extract_ec_field(meta, "deadlineDate"),
                        "conditions": _extract_ec_field(meta, "conditions"),
                    }
    except Exception:
        pass
    return None


# ═══════════════════════════════════════════════════════════
# EURESEARCH SCRAPER
# ═══════════════════════════════════════════════════════════
EURESEARCH_URL = "https://www.euresearch.ch/en/our-services/inform/open-calls-137.html"


def fetch_euresearch_calls(max_results: int = 50) -> List[Dict]:
    if not HAS_BS4:
        return []

    calls = []
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml",
        }
        resp = requests.get(EURESEARCH_URL, headers=headers, timeout=20)
        if resp.status_code != 200:
            return []

        soup = BeautifulSoup(resp.text, "html.parser")
        seen_titles = set()

        for block in soup.find_all(["div", "article", "li", "tr"], recursive=True):
            text = block.get_text(separator=" ", strip=True)
            if len(text) < 30:
                continue
            if not re.search(
                r'(HORIZON|ERC|EIC|MSCA|Pathfinder|Accelerator|'
                r'Starting|Consolidator|Advanced|Doctoral|'
                r'Postdoctoral|Twinning|Teaming|Cluster)',
                text, re.IGNORECASE,
            ):
                continue

            title = ""
            for h in block.find_all(["h2", "h3", "h4", "strong", "b"]):
                t = clean_html(h.get_text(strip=True))
                if len(t) >= 10:
                    title = t
                    break
            if not title:
                lines = [l.strip() for l in text.split("\n") if len(l.strip()) >= 10]
                if lines:
                    title = clean_html(lines[0])[:200]
            if not title or len(title) < 10:
                continue

            tk = title.lower()[:50]
            if tk in seen_titles:
                continue
            seen_titles.add(tk)

            link = ""
            for a in block.find_all("a", href=True):
                href = a.get("href", "")
                if any(x in href.lower() for x in ["topic", "call", "horizon", "erc", "eic", "ec.europa"]):
                    link = href
                    break
            if not link:
                fa = block.find("a", href=True)
                if fa:
                    link = fa.get("href", "")
            if link and not link.startswith("http"):
                link = f"https://www.euresearch.ch{link}"

            status = "Open"
            tl = text.lower()
            if "forthcoming" in tl or "upcoming" in tl:
                status = "Forthcoming"
            elif "closed" in tl:
                status = "Closed"

            deadline = "N/A"
            dm = re.search(r'(\d{4}-\d{2}-\d{2})', text)
            if dm:
                deadline = dm.group(1)
            else:
                dm2 = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
                if dm2:
                    d, m, y = dm2.groups()
                    deadline = f"{y}-{m}-{d}"

            budget = ""
            bm = re.search(r'€\s*[\d.,]+\s*[MmKkBb]?', text)
            if bm:
                budget = bm.group(0).strip()

            action_types = _detect_action_types_from_text(title, text)

            call_id = ""
            im = re.search(
                r'(HORIZON-[A-Z0-9]+-\d{4}-[A-Z0-9-]+|'
                r'ERC-\d{4}-[A-Za-z]+|EIC-\d{4}-[A-Za-z]+|'
                r'MSCA-\d{4}-[A-Za-z-]+)',
                text,
            )
            if im:
                call_id = im.group(1)
            else:
                call_id = f"EUR-{hashlib.md5(title.encode()).hexdigest()[:8]}"

            desc = clean_html(text)
            if desc.startswith(title):
                desc = desc[len(title):].strip()

            calls.append({
                "call_id": call_id,
                "title": title,
                "status": status,
                "deadline": deadline,
                "start_date": "",
                "budget_total": budget,
                "budget_per_project": budget,
                "action_types": action_types,
                "description": desc[:500],
                "keywords": [],
                "destination": "",
                "scope": desc[:1000],
                "expected_outcomes": "",
                "link": link or EURESEARCH_URL,
                "source": "Euresearch",
                "topics": [],
            })

            if len(calls) >= max_results:
                break

    except Exception:
        pass

    return calls


# ═══════════════════════════════════════════════════════════
# UFUK AVRUPA SCRAPER
# ═══════════════════════════════════════════════════════════
UFUKAVRUPA_BASE = "https://ufukavrupa.org.tr"


def fetch_ufukavrupa_calls(max_results: int = 30) -> List[Dict]:
    if not HAS_BS4:
        return []

    calls = []
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 Chrome/120.0",
            "Accept": "text/html",
            "Accept-Language": "tr-TR,tr;q=0.9",
        }

        for url in [f"{UFUKAVRUPA_BASE}/tr/cagrilar", f"{UFUKAVRUPA_BASE}/tr"]:
            try:
                resp = requests.get(url, headers=headers, timeout=20)
                if resp.status_code != 200:
                    continue

                soup = BeautifulSoup(resp.text, "html.parser")
                seen = set()

                for el in soup.find_all(["article", "div", "li"]):
                    text = el.get_text(separator=" ", strip=True)
                    if len(text) < 30:
                        continue
                    if not any(kw in text.lower() for kw in [
                        "çağrı", "horizon", "erc", "eic", "msca",
                        "hibe", "başvuru", "program", "cluster", "call",
                    ]):
                        continue

                    title_el = el.find(["h2", "h3", "h4", "a"])
                    if not title_el:
                        continue
                    title = clean_html(title_el.get_text(strip=True))
                    if not title or len(title) < 10:
                        continue
                    tk = title.lower()[:50]
                    if tk in seen:
                        continue
                    seen.add(tk)

                    link = ""
                    a = title_el if title_el.name == "a" else title_el.find("a")
                    if a:
                        link = a.get("href", "")
                    if link and not link.startswith("http"):
                        link = f"{UFUKAVRUPA_BASE}{link}"

                    deadline = "N/A"
                    dm = re.search(r'(\d{4}-\d{2}-\d{2})', text)
                    if dm:
                        deadline = dm.group(1)

                    calls.append({
                        "call_id": f"UA-{hashlib.md5(title.encode()).hexdigest()[:8]}",
                        "title": title,
                        "status": "Open",
                        "deadline": deadline,
                        "start_date": "",
                        "budget_total": "",
                        "budget_per_project": "",
                        "action_types": _detect_action_types_from_text("", title),
                        "description": clean_html(text)[:500],
                        "keywords": [],
                        "destination": "",
                        "scope": clean_html(text)[:1000],
                        "expected_outcomes": "",
                        "link": link or url,
                        "source": "UfukAvrupa",
                        "topics": [],
                    })
                    if len(calls) >= max_results:
                        break

                if calls:
                    break
            except Exception:
                continue
    except Exception:
        pass

    return calls


# ═══════════════════════════════════════════════════════════
# DETECT ACTION TYPE
# ═══════════════════════════════════════════════════════════
def detect_action_type_from_call(call_data: dict) -> str:
    types = call_data.get("action_types", [])
    title = call_data.get("title", "").lower()
    call_id = call_data.get("call_id", "").lower()
    combined = " ".join(types).lower() + " " + title + " " + call_id

    mapping = [
        (["eic-pathfinder", "eic pathfinder", "pathfinder open"], "EIC-Pathfinder-Open"),
        (["eic-accelerator", "eic accelerator"], "EIC-Accelerator"),
        (["erc-stg", "erc starting", "starting grant"], "ERC-StG"),
        (["erc-cog", "erc consolidator", "consolidator"], "ERC-CoG"),
        (["erc-adg", "erc advanced", "advanced grant"], "ERC-AdG"),
        (["msca-dn", "msca doctoral", "doctoral network"], "MSCA-DN"),
        (["msca-pf", "msca postdoctoral"], "MSCA-PF"),
        (["csa", "coordination and support"], "CSA"),
        (["ia", "innovation action"], "IA"),
        (["ria", "research and innovation"], "RIA"),
    ]
    for keywords, action_type in mapping:
        if any(kw in combined for kw in keywords):
            return action_type
    if types and types[0] != "RIA":
        return types[0]
    return "RIA"


def build_call_specific_criteria(call_data: dict, topic_details: dict = None) -> dict:
    at = detect_action_type_from_call(call_data)
    ctx_parts = [
        f"Call: {clean_html(call_data.get('call_id', 'N/A'))}",
        f"Title: {clean_html(call_data.get('title', 'N/A'))}",
        f"Action Type: {at}",
        f"Deadline: {call_data.get('deadline', 'N/A')}",
    ]
    scope = clean_html(call_data.get("scope", ""))
    if scope:
        ctx_parts.append(f"Scope: {scope[:800]}")
    outcomes = clean_html(call_data.get("expected_outcomes", ""))
    if outcomes:
        ctx_parts.append(f"Expected Outcomes: {outcomes[:500]}")
    if topic_details and topic_details.get("description"):
        ctx_parts.append(f"Topic: {clean_html(topic_details['description'][:1000])}")
    return {
        "action_type": at,
        "evaluation_context": "\n".join(ctx_parts),
        "call_data": call_data,
        "topic_details": topic_details,
    }


# ═══════════════════════════════════════════════════════════
# FETCH ALL
# ═══════════════════════════════════════════════════════════
def fetch_all_calls(
    search_text: str = "",
    status_filter: str = "",
    use_ec_api: bool = True,
    use_euresearch: bool = True,
    use_ufukavrupa: bool = True,
    max_api_results: int = 500,
) -> Tuple[List[Dict], Dict]:
    all_calls = []
    src_stats = {
        "ec_api": 0, "euresearch": 0, "ufukavrupa": 0,
        "local_db": 0, "total": 0, "ec_debug": {},
    }

    if use_ec_api:
        try:
            ec_calls, ec_debug = fetch_horizon_calls(
                search_text, status_filter, max_api_results,
            )
            src_stats["ec_api"] = len(ec_calls)
            src_stats["ec_debug"] = ec_debug
            all_calls.extend(ec_calls)
        except Exception as e:
            src_stats["ec_debug"] = {"error": str(e)[:200]}

    if use_euresearch:
        try:
            eur = fetch_euresearch_calls(50)
            src_stats["euresearch"] = len(eur)
            all_calls.extend(eur)
        except Exception:
            pass

    if use_ufukavrupa:
        try:
            ua = fetch_ufukavrupa_calls(30)
            src_stats["ufukavrupa"] = len(ua)
            all_calls.extend(ua)
        except Exception:
            pass

    try:
        from call_db import HORIZON_CALLS_DB
        existing = set(c.get("call_id", "") for c in all_calls)
        added = 0
        for db_call in HORIZON_CALLS_DB:
            if db_call.get("call_id") not in existing:
                cp = dict(db_call)
                cp["source"] = cp.get("source", "Local DB")
                all_calls.append(cp)
                added += 1
        src_stats["local_db"] = added
    except Exception:
        pass

    seen = set()
    unique = []
    for c in all_calls:
        key = c.get("call_id", "") or c.get("title", "")[:60]
        if key and key not in seen:
            seen.add(key)
            c["title"] = clean_html(c.get("title", ""))
            c["description"] = clean_html(c.get("description", ""))
            c["destination"] = clean_html(c.get("destination", ""))
            unique.append(c)

    src_stats["total"] = len(unique)
    return unique, src_stats


# ═══════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════
def calls_to_excel_bytes(calls: List[Dict]) -> bytes:
    if not HAS_OPENPYXL:
        import csv
        from io import StringIO
        out = StringIO()
        w = csv.writer(out)
        w.writerow(["Call ID", "Title", "Status", "Deadline", "Action Types", "Budget", "Source", "Link"])
        for c in calls:
            w.writerow([
                c.get("call_id", ""), clean_html(c.get("title", "")),
                c.get("status", ""), c.get("deadline", ""),
                ", ".join(c.get("action_types", [])),
                c.get("budget_per_project", c.get("budget_total", "")),
                c.get("source", ""), c.get("link", ""),
            ])
        return out.getvalue().encode("utf-8-sig")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horizon Calls"
    hdrs = ["Call ID", "Title", "Status", "Deadline", "Action Types", "Budget", "Destination", "Source", "Link", "Description"]
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    for row, c in enumerate(calls, 2):
        ws.cell(row=row, column=1, value=c.get("call_id", ""))
        ws.cell(row=row, column=2, value=clean_html(c.get("title", ""))[:200])
        ws.cell(row=row, column=3, value=c.get("status", ""))
        ws.cell(row=row, column=4, value=c.get("deadline", ""))
        ws.cell(row=row, column=5, value=", ".join(c.get("action_types", [])))
        ws.cell(row=row, column=6, value=c.get("budget_per_project", c.get("budget_total", "")))
        ws.cell(row=row, column=7, value=clean_html(c.get("destination", ""))[:100])
        ws.cell(row=row, column=8, value=c.get("source", ""))
        ws.cell(row=row, column=9, value=c.get("link", ""))
        ws.cell(row=row, column=10, value=clean_html(c.get("description", ""))[:300])
    for i, w in enumerate([25, 60, 12, 12, 20, 15, 30, 15, 50, 50], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
